from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import timedelta
from django.core.paginator import Paginator
from accounts.models import User, Role, get_owner_filter
from restaurant.models import Product, MainCategory, SubCategory, TableInfo
from orders.models import Order, OrderItem
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from django.core.exceptions import PermissionDenied
import json
import qrcode
import io
import csv
from decimal import Decimal, InvalidOperation
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import tempfile
import os
try:
    import openpyxl
except ImportError:
    openpyxl = None


def get_production_qr_url(request, qr_code):
    """
    Helper function to generate the correct QR URL
    - Local development: http://127.0.0.1:8000/r/{qr_code}/
    - Production: https://easyfixsoft.com/r/{qr_code}/
    """
    host = request.get_host()
    
    # Force HTTPS for production domains
    if 'easyfixsoft.com' in host or '24.199.116.165' in host:
        return f'https://easyfixsoft.com/r/{qr_code}/'
    
    # Local development - use HTTP
    if '127.0.0.1' in host or 'localhost' in host:
        return f'http://{host}/r/{qr_code}/'
    
    # Fallback - use HTTPS with current host
    return f'https://{host}/r/{qr_code}/'


@login_required
def admin_dashboard(request):
    """Main admin dashboard view - accessible by administrators and owners"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('restaurant:home')

    try:
        owner_filter = get_owner_filter(request.user)
        
        if owner_filter:
            # Owner-specific statistics
            total_users = User.objects.filter(owner=owner_filter).count() + 1  # +1 for the owner themselves
            total_orders = Order.objects.filter(table_info__owner=owner_filter).count()
            total_products = Product.objects.filter(main_category__owner=owner_filter).count()
            total_tables = TableInfo.objects.filter(owner=owner_filter).count()
            
            # Recent orders (last 7 days) for this owner
            seven_days_ago = timezone.now() - timedelta(days=7)
            recent_orders = Order.objects.filter(
                table_info__owner=owner_filter,
                created_at__gte=seven_days_ago
            ).count()
            
            # Today's revenue for this owner
            today = timezone.now().date()
            today_orders = Order.objects.filter(
                table_info__owner=owner_filter,
                created_at__date=today
            )
            today_revenue = today_orders.aggregate(total=Sum('total_amount'))['total'] or 0
            
            # Pending orders for this owner
            pending_orders = Order.objects.filter(
                table_info__owner=owner_filter,
                status='pending'
            ).count()
        else:
            # Administrator sees all statistics
            total_users = User.objects.count()
            total_orders = Order.objects.count()
            total_products = Product.objects.count()
            total_tables = TableInfo.objects.count()
            
            # Recent orders (last 7 days)
            seven_days_ago = timezone.now() - timedelta(days=7)
            recent_orders = Order.objects.filter(created_at__gte=seven_days_ago).count()
            
            # Today's revenue
            today = timezone.now().date()
            today_orders = Order.objects.filter(created_at__date=today)
            today_revenue = today_orders.aggregate(total=Sum('total_amount'))['total'] or 0
            
            # Pending orders
            pending_orders = Order.objects.filter(status='pending').count()
            
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    context = {
        'total_users': total_users,
        'total_orders': total_orders,
        'total_products': total_products,
        'total_tables': total_tables,
        'recent_orders': recent_orders,
        'today_revenue': today_revenue,
        'pending_orders': pending_orders,
        'restaurant_name': request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants",
    }

    return render(request, 'admin_panel/dashboard.html', context)


@login_required
def manage_users(request):
    """User management view"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator/Owner privileges required.")
        return redirect('restaurant:home')

    try:
        owner_filter = get_owner_filter(request.user)
        
        if owner_filter:
            # Owners see only their own staff and customers (full access to all roles except administrator)
            users = User.objects.filter(
                owner=owner_filter
            ).order_by('-date_joined')
            roles = Role.objects.exclude(name='administrator')
        else:
            # Administrators see all users and roles
            users = User.objects.all().order_by('-date_joined')
            roles = Role.objects.all()
            
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    context = {
        'users': users,
        'roles': roles,
        'is_owner_access': request.user.is_owner() and not request.user.is_administrator(),
        'restaurant_name': request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants",
    }

    return render(request, 'admin_panel/manage_users.html', context)


@login_required
def manage_products(request):
    """Product management view - accessible by administrators and owners"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('restaurant:home')

    try:
        owner_filter = get_owner_filter(request.user)
        
        if owner_filter:
            # Owner sees only their own products and categories
            main_categories = MainCategory.objects.filter(
                is_active=True, 
                owner=owner_filter
            ).order_by('name')
            
            all_products = Product.objects.filter(
                main_category__owner=owner_filter
            ).select_related('main_category', 'sub_category').order_by('name')
        else:
            # Administrator sees all products and categories
            main_categories = MainCategory.objects.filter(is_active=True).order_by('name')
            
            all_products = Product.objects.select_related('main_category', 'sub_category').all().order_by('name')
        
        # Add pagination for each category
        paginated_categories = []
        for category in main_categories:
            category_products = all_products.filter(main_category=category)
            
            # Get page number for this category (default 1)
            page_param = f'page_{category.id}'
            page_number = request.GET.get(page_param, 1)
            
            # Paginate products (10 per page)
            paginator = Paginator(category_products, 10)
            page_obj = paginator.get_page(page_number)
            
            # Only include categories that have products
            if category_products.exists():
                paginated_categories.append({
                    'category': category,
                    'products': page_obj,
                    'page_param': page_param,
                    'total_count': category_products.count()
                })
            
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    context = {
        'main_categories': paginated_categories,
        'all_main_categories': main_categories,  # Add this for the form dropdown
        'restaurant_name': request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants",
    }

    return render(request, 'admin_panel/manage_products.html', context)


@login_required
def manage_orders(request):
    """Order management view with status-based tabs - accessible by administrators and owners"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('restaurant:home')

    try:
        owner_filter = get_owner_filter(request.user)
        
        if owner_filter:
            # Owner sees only orders placed at their restaurant's tables
            base_orders = Order.objects.filter(table_info__owner=owner_filter)
        else:
            # Administrator sees all orders
            base_orders = Order.objects.all()
            
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    # Organize orders by status with counts
    pending_orders = base_orders.filter(status='pending').order_by('-created_at')
    confirmed_orders = base_orders.filter(status='confirmed').order_by('-created_at')
    preparing_orders = base_orders.filter(status='preparing').order_by('-created_at')
    ready_orders = base_orders.filter(status='ready').order_by('-created_at')
    served_orders = base_orders.filter(status='served').order_by('-created_at')
    cancelled_orders = base_orders.filter(status='cancelled').order_by('-created_at')
    
    # Get counts for tab badges
    pending_count = pending_orders.count()
    confirmed_count = confirmed_orders.count()
    preparing_count = preparing_orders.count()
    ready_count = ready_orders.count()
    served_count = served_orders.count()
    cancelled_count = cancelled_orders.count()
    total_count = base_orders.count()

    context = {
        'pending_orders': pending_orders,
        'confirmed_orders': confirmed_orders,
        'preparing_orders': preparing_orders,
        'ready_orders': ready_orders,
        'served_orders': served_orders,
        'cancelled_orders': cancelled_orders,
        'pending_count': pending_count,
        'confirmed_count': confirmed_count,
        'preparing_count': preparing_count,
        'ready_count': ready_count,
        'served_count': served_count,
        'cancelled_count': cancelled_count,
        'total_count': total_count,
        'restaurant_name': request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants",
    }

    return render(request, 'admin_panel/manage_orders.html', context)


@login_required
def manage_tables(request):
    """Table management view"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator/Owner privileges required.")
        return redirect('restaurant:home')

    try:
        owner_filter = get_owner_filter(request.user)
        
        if owner_filter:
            # Owner sees only their own tables
            tables = TableInfo.objects.filter(owner=owner_filter)
        else:
            # Administrator sees all tables
            tables = TableInfo.objects.all()
            
        # Custom sorting to handle T01, T02, T10, T011 properly
        tables = sorted(tables, key=lambda x: (len(x.tbl_no), x.tbl_no))
        
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    context = {
        'tables': tables,
        'restaurant_name': request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants",
    }

    return render(request, 'admin_panel/manage_tables.html', context)


@login_required
def manage_categories(request):
    """Category management view"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator/Owner privileges required.")
        return redirect('restaurant:home')

    try:
        owner_filter = get_owner_filter(request.user)
        
        if owner_filter:
            # Owner sees only their own categories
            main_categories = MainCategory.objects.filter(owner=owner_filter).order_by('name')
            subcategories = SubCategory.objects.filter(
                main_category__owner=owner_filter
            ).order_by('main_category__name', 'name')
        else:
            # Administrator sees all categories
            main_categories = MainCategory.objects.all().order_by('name')
            subcategories = SubCategory.objects.all().order_by('main_category__name', 'name')
            
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    context = {
        'main_categories': main_categories,
        'subcategories': subcategories,
        'restaurant_name': request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants",
    }

    return render(request, 'admin_panel/manage_categories.html', context)


@login_required
@require_POST
def add_main_category(request):
    """Add a new main category"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        image = request.FILES.get('image')  # Handle image upload

        if not name:
            return JsonResponse({'success': False, 'message': 'Category name is required'})

        # Check for duplicate names within the same owner's categories
        if owner_filter:
            if MainCategory.objects.filter(name__iexact=name, owner=owner_filter).exists():
                return JsonResponse({'success': False, 'message': 'Category with this name already exists in your restaurant'})
        else:
            if MainCategory.objects.filter(name__iexact=name).exists():
                return JsonResponse({'success': False, 'message': 'Category with this name already exists'})

        category = MainCategory.objects.create(
            name=name,
            description=description,
            image=image,  # Add image to creation
            owner=owner_filter if owner_filter else None
        )

        return JsonResponse({
            'success': True,
            'message': 'Main category added successfully',
            'category': {
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'is_active': category.is_active,
                'image_url': category.image.url if category.image else None
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def edit_main_category(request, category_id):
    """Edit an existing main category"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get the category and check owner permission
        if owner_filter:
            category = get_object_or_404(MainCategory, id=category_id, owner=owner_filter)
        else:
            category = get_object_or_404(MainCategory, id=category_id)
            
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        image = request.FILES.get('image')  # Handle image upload

        if not name:
            return JsonResponse({'success': False, 'message': 'Category name is required'})

        # Check for duplicate names within the same owner's categories
        if owner_filter:
            if MainCategory.objects.filter(name__iexact=name, owner=owner_filter).exclude(id=category_id).exists():
                return JsonResponse({'success': False, 'message': 'Category with this name already exists in your restaurant'})
        else:
            if MainCategory.objects.filter(name__iexact=name).exclude(id=category_id).exists():
                return JsonResponse({'success': False, 'message': 'Category with this name already exists'})

        category.name = name
        category.description = description
        if image:  # Only update image if a new one is uploaded
            category.image = image
        category.save()

        return JsonResponse({
            'success': True,
            'message': 'Main category updated successfully',
            'category': {
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'is_active': category.is_active,
                'image_url': category.image.url if category.image else None
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def delete_main_category(request, category_id):
    """Delete a main category"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get category with owner filtering
        if owner_filter:
            category = get_object_or_404(MainCategory, id=category_id, owner=owner_filter)
        else:
            category = get_object_or_404(MainCategory, id=category_id)
        category_name = category.name
        subcategory_count = category.subcategories.count()
        
        # Delete the category (this will cascade delete subcategories and products due to ON DELETE CASCADE)
        category.delete()

        if subcategory_count > 0:
            message = f'Main category "{category_name}" and its {subcategory_count} subcategories deleted successfully'
        else:
            message = f'Main category "{category_name}" deleted successfully'

        return JsonResponse({
            'success': True,
            'message': message
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def toggle_main_category(request, category_id):
    """Toggle main category active status"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get category with owner filtering
        if owner_filter:
            category = get_object_or_404(MainCategory, id=category_id, owner=owner_filter)
        else:
            category = get_object_or_404(MainCategory, id=category_id)
        category.is_active = not category.is_active
        category.save()

        status = 'activated' if category.is_active else 'deactivated'

        return JsonResponse({
            'success': True,
            'message': f'Main category "{category.name}" {status} successfully',
            'is_active': category.is_active
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def add_subcategory(request):
    """Add a new subcategory"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        main_category_id = request.POST.get('main_category')
        name = request.POST.get('name')
        description = request.POST.get('description', '')

        if not main_category_id or not name:
            return JsonResponse({'success': False, 'message': 'Main category and subcategory name are required'})

        # Get main category with owner filtering
        if owner_filter:
            try:
                main_category = MainCategory.objects.get(id=main_category_id, owner=owner_filter)
            except MainCategory.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Main category not found or access denied'})
        else:
            try:
                main_category = MainCategory.objects.get(id=main_category_id)
            except MainCategory.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Main category not found'})

        if SubCategory.objects.filter(main_category=main_category, name__iexact=name).exists():
            return JsonResponse({'success': False, 'message': 'Subcategory with this name already exists in the selected main category'})

        subcategory = SubCategory.objects.create(
            main_category=main_category,
            name=name,
            description=description
        )

        return JsonResponse({
            'success': True,
            'message': 'Subcategory added successfully',
            'subcategory': {
                'id': subcategory.id,
                'name': subcategory.name,
                'description': subcategory.description,
                'main_category': subcategory.main_category.name,
                'is_active': subcategory.is_active
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def edit_subcategory(request, subcategory_id):
    """Edit an existing subcategory"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get subcategory with owner filtering
        if owner_filter:
            subcategory = get_object_or_404(SubCategory, id=subcategory_id, main_category__owner=owner_filter)
        else:
            subcategory = get_object_or_404(SubCategory, id=subcategory_id)
            
        main_category_id = request.POST.get('main_category')
        name = request.POST.get('name')
        description = request.POST.get('description', '')

        if not main_category_id or not name:
            return JsonResponse({'success': False, 'message': 'Main category and subcategory name are required'})

        # Get main category with owner filtering
        if owner_filter:
            main_category = get_object_or_404(MainCategory, id=main_category_id, owner=owner_filter)
        else:
            main_category = get_object_or_404(MainCategory, id=main_category_id)

        if SubCategory.objects.filter(main_category=main_category, name__iexact=name).exclude(id=subcategory_id).exists():
            return JsonResponse({'success': False, 'message': 'Subcategory with this name already exists in the selected main category'})

        subcategory.main_category = main_category
        subcategory.name = name
        subcategory.description = description
        subcategory.save()

        return JsonResponse({
            'success': True,
            'message': 'Subcategory updated successfully',
            'subcategory': {
                'id': subcategory.id,
                'name': subcategory.name,
                'description': subcategory.description,
                'main_category': subcategory.main_category.name,
                'is_active': subcategory.is_active
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def delete_subcategory(request, subcategory_id):
    """Delete a subcategory"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get subcategory with owner filtering
        if owner_filter:
            subcategory = get_object_or_404(SubCategory, id=subcategory_id, main_category__owner=owner_filter)
        else:
            subcategory = get_object_or_404(SubCategory, id=subcategory_id)
        subcategory_name = subcategory.name
        product_count = subcategory.products.count()
        
        # Delete the subcategory (this will cascade delete products due to ON DELETE CASCADE)
        subcategory.delete()

        if product_count > 0:
            message = f'Subcategory "{subcategory_name}" and its {product_count} products deleted successfully'
        else:
            message = f'Subcategory "{subcategory_name}" deleted successfully'

        return JsonResponse({
            'success': True,
            'message': message
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def toggle_subcategory(request, subcategory_id):
    """Toggle subcategory active status"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get subcategory with owner filtering
        if owner_filter:
            subcategory = get_object_or_404(SubCategory, id=subcategory_id, main_category__owner=owner_filter)
        else:
            subcategory = get_object_or_404(SubCategory, id=subcategory_id)
        subcategory.is_active = not subcategory.is_active
        subcategory.save()

        status = 'activated' if subcategory.is_active else 'deactivated'

        return JsonResponse({
            'success': True,
            'message': f'Subcategory "{subcategory.name}" {status} successfully',
            'is_active': subcategory.is_active
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# Product CRUD API Views
@login_required
def get_subcategories(request, main_category_id):
    """Get subcategories for a main category"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        if owner_filter:
            main_category = get_object_or_404(MainCategory, id=main_category_id, owner=owner_filter)
        else:
            main_category = get_object_or_404(MainCategory, id=main_category_id)
            
        subcategories = main_category.subcategories.filter(is_active=True).values('id', 'name')
        
        return JsonResponse({
            'success': True,
            'subcategories': list(subcategories)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def add_product(request):
    """Add new product"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get form data
        name = request.POST.get('name')
        description = request.POST.get('description')
        main_category_id = request.POST.get('main_category')
        sub_category_id = request.POST.get('sub_category')
        price = request.POST.get('price')
        stock = request.POST.get('available_in_stock')
        prep_time = request.POST.get('preparation_time', 15)
        is_available = request.POST.get('is_available') == 'on'
        image = request.FILES.get('image')
        
        # Validate required fields
        if not all([name, description, main_category_id, sub_category_id, price, stock]):
            return JsonResponse({'success': False, 'message': 'All required fields must be filled'})
        
        # Get category objects with owner filtering
        if owner_filter:
            main_category = get_object_or_404(MainCategory, id=main_category_id, owner=owner_filter)
            sub_category = get_object_or_404(SubCategory, id=sub_category_id, main_category__owner=owner_filter)
        else:
            main_category = get_object_or_404(MainCategory, id=main_category_id)
            sub_category = get_object_or_404(SubCategory, id=sub_category_id)
        
        # Create product
        product = Product.objects.create(
            name=name,
            description=description,
            main_category=main_category,
            sub_category=sub_category,
            price=float(price),
            available_in_stock=int(stock),
            preparation_time=int(prep_time),
            is_available=is_available,
            image=image
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Product "{product.name}" added successfully',
            'product_id': product.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def view_product(request, product_id):
    """Get product details for viewing"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get product with owner filtering
        if owner_filter:
            product = get_object_or_404(Product, id=product_id, main_category__owner=owner_filter)
        else:
            product = get_object_or_404(Product, id=product_id)
        
        html = render_to_string('admin_panel/product_detail.html', {
            'product': product
        })
        
        return JsonResponse({
            'success': True,
            'html': html
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def edit_product(request, product_id):
    """Get product details for editing"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get product with owner filtering
        if owner_filter:
            product = get_object_or_404(Product, id=product_id, main_category__owner=owner_filter)
            main_categories = MainCategory.objects.filter(is_active=True, owner=owner_filter).order_by('name')
        else:
            product = get_object_or_404(Product, id=product_id)
            main_categories = MainCategory.objects.filter(is_active=True).order_by('name')
        subcategories = product.main_category.subcategories.filter(is_active=True).order_by('name')
        
        html = render_to_string('admin_panel/product_edit_form.html', {
            'product': product,
            'main_categories': main_categories,
            'subcategories': subcategories
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'html': html
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def update_product(request, product_id):
    """Update product"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Access denied'})
        else:
            messages.error(request, "Access denied. Administrator privileges required.")
            return redirect('admin_panel:manage_products')
    
    try:
        product = get_object_or_404(Product, id=product_id)
        
        # Update fields
        product.name = request.POST.get('name', product.name)
        product.description = request.POST.get('description', product.description)
        
        if request.POST.get('main_category'):
            product.main_category = get_object_or_404(MainCategory, id=request.POST.get('main_category'))
        
        if request.POST.get('sub_category'):
            product.sub_category = get_object_or_404(SubCategory, id=request.POST.get('sub_category'))
        
        if request.POST.get('price'):
            product.price = float(request.POST.get('price'))
        
        if request.POST.get('available_in_stock'):
            product.available_in_stock = int(request.POST.get('available_in_stock'))
        
        if request.POST.get('preparation_time'):
            product.preparation_time = int(request.POST.get('preparation_time'))
        
        product.is_available = request.POST.get('is_available') == 'on'
        
        product.save()
        
        # Check if this is an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Product "{product.name}" updated successfully'
            })
        else:
            # Regular form submission - redirect with success message
            messages.success(request, f'Product "{product.name}" updated successfully')
            return redirect('admin_panel:manage_products')
        
    except Exception as e:
        print(f"Error updating product: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': str(e)})
        else:
            messages.error(request, f'Error updating product: {str(e)}')
            return redirect('admin_panel:manage_products')


@login_required
@require_http_methods(["POST"])
def toggle_product_availability(request, product_id):
    """Toggle product availability"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        product = get_object_or_404(Product, id=product_id)
        data = json.loads(request.body)
        
        product.is_available = data.get('is_available', not product.is_available)
        product.save()
        
        status = 'available' if product.is_available else 'unavailable'
        
        return JsonResponse({
            'success': True,
            'message': f'Product "{product.name}" is now {status}',
            'is_available': product.is_available
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def delete_product(request, product_id):
    """Delete product"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        product = get_object_or_404(Product, id=product_id)
        product_name = product.name
        product.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Product "{product_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ============================================================================
# USER MANAGEMENT CRUD OPERATIONS
# ============================================================================

@login_required
@require_POST
def add_user(request):
    """Add a new user"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Get form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        role_name = request.POST.get('role', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        phone_number = request.POST.get('phone_number', '').strip()
        address = request.POST.get('address', '').strip()
        
        # Validation
        if not all([first_name, last_name, username, email, password, role_name]):
            return JsonResponse({'success': False, 'message': 'All required fields must be filled'})
        
        # Check role restrictions for owners
        if request.user.is_owner() and not request.user.is_administrator():
            # Owners can create any role except administrator
            if role_name == 'administrator':
                return JsonResponse({'success': False, 'message': 'Owners cannot create administrator accounts'})
        
        # Check if username or email already exists
        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'message': 'Username already exists'})
        
        if User.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message': 'Email already exists'})
        
        # Get role
        try:
            role = Role.objects.get(name=role_name)
        except Role.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid role selected'})
        
        # Create user
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            role=role,
            is_active=is_active,
            phone_number=phone_number,
            address=address
        )
        
        # Assign owner for non-administrator users
        if request.user.is_owner() and not request.user.is_administrator():
            # Owners assign their staff to themselves
            user.owner = request.user
            user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'User "{user.get_full_name() or user.username}" created successfully',
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'email': user.email,
                'role': role.get_name_display(),
                'is_active': user.is_active,
                'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M')
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error creating user: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def get_user_data(request, user_id):
    """Get user data for editing"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if owner is trying to access a user they're not allowed to manage
        if request.user.is_owner() and not request.user.is_administrator():
            # Owners can manage any user except administrators
            if user.role and user.role.name == 'administrator':
                return JsonResponse({'success': False, 'message': 'Access denied - you cannot manage administrator accounts'})
        
        return JsonResponse({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'role': user.role.name if user.role else '',
                'is_active': user.is_active,
                'phone_number': user.phone_number,
                'address': user.address,
                'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M'),
                'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never'
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def update_user(request, user_id):
    """Update an existing user"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if owner is trying to update a user they're not allowed to manage
        if request.user.is_owner() and not request.user.is_administrator():
            # Owners can manage any user except administrators
            if user.role and user.role.name == 'administrator':
                return JsonResponse({'success': False, 'message': 'Access denied - you cannot manage administrator accounts'})
        
        # Get form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        role_name = request.POST.get('role', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        phone_number = request.POST.get('phone_number', '').strip()
        address = request.POST.get('address', '').strip()
        
        # Validation
        if not all([first_name, last_name, username, email, role_name]):
            return JsonResponse({'success': False, 'message': 'All required fields must be filled'})
        
        # Check if username or email already exists (excluding current user)
        if User.objects.filter(username=username).exclude(id=user_id).exists():
            return JsonResponse({'success': False, 'message': 'Username already exists'})
        
        if User.objects.filter(email=email).exclude(id=user_id).exists():
            return JsonResponse({'success': False, 'message': 'Email already exists'})
        
        # Get role
        try:
            role = Role.objects.get(name=role_name)
        except Role.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid role selected'})
        
        # Check role restrictions for owners
        if request.user.is_owner() and not request.user.is_administrator():
            # Owners can assign any role except administrator
            if role_name == 'administrator':
                return JsonResponse({'success': False, 'message': 'Owners cannot assign administrator role'})
        
        # Update user
        user.first_name = first_name
        user.last_name = last_name
        user.username = username
        user.email = email
        user.role = role
        user.is_active = is_active
        user.phone_number = phone_number
        user.address = address
        
        # Update password if provided
        if password:
            user.set_password(password)
        
        user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'User "{user.get_full_name() or user.username}" updated successfully',
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'email': user.email,
                'role': role.get_name_display(),
                'is_active': user.is_active
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error updating user: {str(e)}'})


@login_required
@require_POST
def toggle_user_status(request, user_id):
    """Toggle user active/inactive status"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if owner is trying to manage a user they're not allowed to
        if request.user.is_owner() and not request.user.is_administrator():
            # Owners can manage any user except administrators
            if user.role and user.role.name == 'administrator':
                return JsonResponse({'success': False, 'message': 'Access denied - you cannot manage administrator accounts'})
        
        # Prevent deactivating self
        if user.id == request.user.id:
            return JsonResponse({'success': False, 'message': 'You cannot deactivate your own account'})
        
        # Prevent deactivating other administrators unless you're a superuser
        if user.is_administrator() and not request.user.is_superuser:
            return JsonResponse({'success': False, 'message': 'You cannot modify administrator accounts'})
        
        user.is_active = not user.is_active
        user.save()
        
        status = 'activated' if user.is_active else 'deactivated'
        
        return JsonResponse({
            'success': True,
            'message': f'User "{user.get_full_name() or user.username}" {status} successfully',
            'is_active': user.is_active
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def delete_user(request, user_id):
    """Delete a user"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if owner is trying to delete a user they're not allowed to manage
        if request.user.is_owner() and not request.user.is_administrator():
            # Owners can delete any user except administrators
            if user.role and user.role.name == 'administrator':
                return JsonResponse({'success': False, 'message': 'Access denied - you cannot delete administrator accounts'})
        
        # Prevent deleting self
        if user.id == request.user.id:
            return JsonResponse({'success': False, 'message': 'You cannot delete your own account'})
        
        # Prevent deleting other administrators unless you're a superuser
        if user.is_administrator() and not request.user.is_superuser:
            return JsonResponse({'success': False, 'message': 'You cannot delete administrator accounts'})
        
        user_name = user.get_full_name() or user.username
        user.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'User "{user_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ============================================================================
# ROLE MANAGEMENT CRUD OPERATIONS
# ============================================================================

@login_required
@require_POST
def add_role(request):
    """Add a new role"""
    if not request.user.is_administrator():
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        
        # Validation
        if not name:
            return JsonResponse({'success': False, 'message': 'Role name is required'})
        
        # Check if role already exists
        if Role.objects.filter(name=name).exists():
            return JsonResponse({'success': False, 'message': 'Role already exists'})
        
        # Validate role name is in choices
        valid_roles = [choice[0] for choice in Role.ROLE_CHOICES]
        if name not in valid_roles:
            return JsonResponse({'success': False, 'message': 'Invalid role name'})
        
        # Create role
        role = Role.objects.create(
            name=name,
            description=description
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Role "{role.get_name_display()}" created successfully',
            'role': {
                'id': role.id,
                'name': role.name,
                'display_name': role.get_name_display(),
                'description': role.description,
                'created_at': role.created_at.strftime('%Y-%m-%d %H:%M')
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error creating role: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def get_role_data(request, role_id):
    """Get role data for editing"""
    if not request.user.is_administrator():
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        role = get_object_or_404(Role, id=role_id)
        
        return JsonResponse({
            'success': True,
            'role': {
                'id': role.id,
                'name': role.name,
                'display_name': role.get_name_display(),
                'description': role.description,
                'created_at': role.created_at.strftime('%Y-%m-%d %H:%M'),
                'user_count': role.user_set.count()
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def update_role(request, role_id):
    """Update an existing role"""
    if not request.user.is_administrator():
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        role = get_object_or_404(Role, id=role_id)
        
        description = request.POST.get('description', '').strip()
        
        # Update role (name cannot be changed for system roles)
        role.description = description
        role.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Role "{role.get_name_display()}" updated successfully',
            'role': {
                'id': role.id,
                'name': role.name,
                'display_name': role.get_name_display(),
                'description': role.description
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error updating role: {str(e)}'})


@login_required
@require_POST
def delete_role(request, role_id):
    """Delete a role"""
    if not request.user.is_administrator():
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        role = get_object_or_404(Role, id=role_id)
        
        # Check if role has users assigned
        if role.user_set.exists():
            return JsonResponse({
                'success': False, 
                'message': f'Cannot delete role "{role.get_name_display()}" because it has users assigned to it'
            })
        
        # Prevent deleting system roles
        system_roles = ['administrator', 'owner', 'customer_care', 'kitchen', 'customer']
        if role.name in system_roles:
            return JsonResponse({
                'success': False, 
                'message': f'Cannot delete system role "{role.get_name_display()}"'
            })
        
        role_name = role.get_name_display()
        role.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Role "{role_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def edit_user(request, user_id):
    """Edit user view (GET request)"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator/Owner privileges required.")
        return redirect('admin_panel:manage_users')
    
    user = get_object_or_404(User, id=user_id)
    
    # Check if owner is trying to edit a user they're not allowed to manage
    if request.user.is_owner() and not request.user.is_administrator():
        # Owners can manage any user except administrators
        if user.role and user.role.name == 'administrator':
            messages.error(request, "Access denied - you cannot manage administrator accounts.")
            return redirect('admin_panel:manage_users')
    
    # Filter roles based on user permissions
    if request.user.is_owner() and not request.user.is_administrator():
        # Owners can assign any role except administrator
        roles = Role.objects.exclude(name='administrator')
    else:
        roles = Role.objects.all()
    
    context = {
        'edit_user': user,
        'roles': roles,
    }
    
    return render(request, 'admin_panel/edit_user.html', context)


@login_required
def edit_role(request, role_id):
    """Edit role view (GET request)"""
    if not request.user.is_administrator():
        messages.error(request, "Access denied. Administrator privileges required.")
        return redirect('admin_panel:manage_users')
    
    role = get_object_or_404(Role, id=role_id)
    
    context = {
        'edit_role': role,
    }
    
    return render(request, 'admin_panel/edit_role.html', context)


# Table Management CRUD Views
@login_required
@require_http_methods(["POST"])
def add_table(request):
    """Add new table"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        table_number = request.POST.get('tbl_no', '').strip()
        capacity = request.POST.get('capacity')
        is_available = request.POST.get('is_available') == 'on'
        
        if not table_number:
            return JsonResponse({'success': False, 'message': 'Table number is required'})
        
        # Basic validation for table number (alphanumeric, max 10 chars)
        if len(table_number) > 10:
            return JsonResponse({'success': False, 'message': 'Table number must be 10 characters or less'})
        
        if not table_number.replace(' ', '').replace('-', '').isalnum():
            return JsonResponse({'success': False, 'message': 'Table number can only contain letters, numbers, spaces, and hyphens'})
        
        if not capacity or int(capacity) < 1:
            return JsonResponse({'success': False, 'message': 'Valid capacity is required'})
        
        # Check if table number already exists for this owner
        if owner_filter:
            if TableInfo.objects.filter(tbl_no=table_number, owner=owner_filter).exists():
                return JsonResponse({'success': False, 'message': 'Table number already exists in your restaurant'})
        else:
            if TableInfo.objects.filter(tbl_no=table_number).exists():
                return JsonResponse({'success': False, 'message': 'Table number already exists'})
        
        # Create new table
        table = TableInfo.objects.create(
            tbl_no=table_number,
            capacity=int(capacity),
            is_available=is_available,
            owner=owner_filter if owner_filter else None
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Table {table_number} added successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def get_table(request):
    """Get table data for editing"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    table_id = request.GET.get('table_id')
    try:
        table = get_object_or_404(TableInfo, id=table_id)
        return JsonResponse({
            'success': True,
            'table': {
                'id': table.id,
                'tbl_no': table.tbl_no,
                'capacity': table.capacity,
                'is_available': table.is_available
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def update_table(request):
    """Update table"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        table_id = request.POST.get('table_id')
        table_number = request.POST.get('tbl_no', '').strip()
        capacity = request.POST.get('capacity')
        is_available = request.POST.get('is_available') == 'on'
        
        if not table_number:
            return JsonResponse({'success': False, 'message': 'Table number is required'})
        
        # Basic validation for table number (alphanumeric, max 10 chars)
        if len(table_number) > 10:
            return JsonResponse({'success': False, 'message': 'Table number must be 10 characters or less'})
        
        if not table_number.replace(' ', '').replace('-', '').isalnum():
            return JsonResponse({'success': False, 'message': 'Table number can only contain letters, numbers, spaces, and hyphens'})
        
        if not capacity or int(capacity) < 1:
            return JsonResponse({'success': False, 'message': 'Valid capacity is required'})
        
        # Get table with owner filtering
        if owner_filter:
            table = get_object_or_404(TableInfo, id=table_id, owner=owner_filter)
        else:
            table = get_object_or_404(TableInfo, id=table_id)
        
        # Check if table number already exists within owner's restaurant (excluding current table)
        if owner_filter:
            if TableInfo.objects.filter(tbl_no=table_number, owner=owner_filter).exclude(id=table_id).exists():
                return JsonResponse({'success': False, 'message': 'Table number already exists in your restaurant'})
        else:
            if TableInfo.objects.filter(tbl_no=table_number).exclude(id=table_id).exists():
                return JsonResponse({'success': False, 'message': 'Table number already exists'})
        
        # Update table
        table.tbl_no = table_number
        table.capacity = int(capacity)
        table.is_available = is_available
        table.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Table {table_number} updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def toggle_table_status(request):
    """Toggle table availability status"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        table_id = request.POST.get('table_id')
        action = request.POST.get('action')  # 'occupy' or 'free'
        
        table = get_object_or_404(TableInfo, id=table_id)
        
        if action == 'occupy':
            table.is_available = False
            message = f'Table {table.tbl_no} marked as occupied'
        elif action == 'free':
            table.is_available = True
            message = f'Table {table.tbl_no} marked as available'
        else:
            return JsonResponse({'success': False, 'message': 'Invalid action'})
        
        table.save()
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def delete_table(request):
    """Delete table"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        table_id = request.POST.get('table_id')
        table = get_object_or_404(TableInfo, id=table_id)
        
        table_number = table.tbl_no
        table.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Table {table_number} deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# Order Management CRUD Views
@login_required
def view_order(request, order_id):
    """View order details"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        order = get_object_or_404(Order, id=order_id)
        order_items = order.order_items.all()
        
        context = {
            'order': order,
            'order_items': order_items,
        }
        
        html = render_to_string('admin_panel/order_details.html', context, request=request)
        return JsonResponse({'success': True, 'html': html})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def update_order_status(request):
    """Update order status"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('status')
        
        if not order_id or not new_status:
            return JsonResponse({'success': False, 'message': 'Order ID and status are required'})
        
        order = get_object_or_404(Order, id=order_id)
        
        # Validate status
        valid_statuses = [choice[0] for choice in Order.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'message': 'Invalid status'})
        
        old_status = order.status
        order.status = new_status
        
        # If confirming order, set confirmed_by
        if new_status == 'confirmed' and not order.confirmed_by:
            order.confirmed_by = request.user
        
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Order #{order.order_number} status updated from {old_status} to {new_status}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def add_order(request):
    """Add new order"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator privileges required.")
        return redirect('restaurant:home')
    
    if request.method == 'POST':
        try:
            owner_filter = get_owner_filter(request.user)
            
            table_id = request.POST.get('table_id')
            customer_id = request.POST.get('customer_id')
            special_instructions = request.POST.get('special_instructions', '')
            
            if not table_id or not customer_id:
                return JsonResponse({'success': False, 'message': 'Table and customer are required'})
            
            # Get table and customer with owner filtering
            if owner_filter:
                table = get_object_or_404(TableInfo, id=table_id, owner=owner_filter)
                customer = get_object_or_404(User, id=customer_id, owner=owner_filter)
            else:
                table = get_object_or_404(TableInfo, id=table_id)
                customer = get_object_or_404(User, id=customer_id)
            
            # Generate order number
            import random
            import string
            order_number = ''.join(random.choices(string.digits, k=8))
            while Order.objects.filter(order_number=order_number).exists():
                order_number = ''.join(random.choices(string.digits, k=8))
            
            # Create order
            order = Order.objects.create(
                order_number=order_number,
                table_info=table,
                ordered_by=customer,
                special_instructions=special_instructions,
                status='pending'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Order #{order_number} created successfully'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    # GET request - show form with owner filtering
    owner_filter = get_owner_filter(request.user)
    
    if owner_filter:
        tables = TableInfo.objects.filter(is_available=True, owner=owner_filter)
        # Get customers belonging to this owner
        role_customer = Role.objects.get(name='customer')
        customers = User.objects.filter(role=role_customer, owner=owner_filter)
    else:
        tables = TableInfo.objects.filter(is_available=True)
        role_customer = Role.objects.get(name='customer')
        customers = User.objects.filter(role=role_customer)
    
    context = {
        'tables': tables,
        'customers': customers,
    }
    
    return render(request, 'admin_panel/add_order.html', context)


@login_required
def edit_order(request, order_id):
    """Edit order"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator privileges required.")
        return redirect('restaurant:home')
    
    # Get owner filter for multi-tenant support
    owner_filter = get_owner_filter(request.user)
    
    # Filter order by owner if owner, or get any order if administrator
    if request.user.is_owner():
        order = get_object_or_404(Order, id=order_id, table_info__owner=owner_filter)
    else:
        order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        try:
            table_id = request.POST.get('table_id')
            customer_id = request.POST.get('customer_id')
            special_instructions = request.POST.get('special_instructions', '')
            status = request.POST.get('status')
            
            if table_id:
                # Filter table by owner
                if request.user.is_owner():
                    table = get_object_or_404(TableInfo, id=table_id, owner=owner_filter)
                else:
                    table = get_object_or_404(TableInfo, id=table_id)
                order.table_info = table
            
            if customer_id:
                order.ordered_by = get_object_or_404(User, id=customer_id)
            
            order.special_instructions = special_instructions
            
            if status and status in [choice[0] for choice in Order.STATUS_CHOICES]:
                order.status = status
            
            order.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Order #{order.order_number} updated successfully'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    # GET request - show form
    # Filter tables and customers by owner if owner, otherwise show all
    if request.user.is_owner():
        tables = TableInfo.objects.filter(owner=owner_filter)
        customers = User.objects.filter(role__name='Customer')  # Customers can be from any restaurant
    else:
        tables = TableInfo.objects.all()
        customers = User.objects.filter(role__name='Customer')
    
    context = {
        'order': order,
        'tables': tables,
        'customers': customers,
        'status_choices': Order.STATUS_CHOICES,
    }
    
    # Check if this is an AJAX request for modal content
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'admin_panel/edit_order_modal.html', context)
    else:
        return render(request, 'admin_panel/edit_order.html', context)


@login_required
@require_http_methods(["POST"])
def delete_order(request, order_id):
    """Delete order"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        order = get_object_or_404(Order, id=order_id)
        order_number = order.order_number
        
        # Only allow deletion of pending or cancelled orders
        if order.status not in ['pending', 'cancelled']:
            return JsonResponse({
                'success': False, 
                'message': 'Only pending or cancelled orders can be deleted'
            })
        
        order.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Order #{order_number} deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def profile(request):
    """User profile view - accessible by all admin panel users"""
    if not (request.user.is_administrator() or request.user.is_owner() or request.user.is_kitchen_staff() or request.user.is_customer_care()):
        messages.error(request, "Access denied. Admin panel access required.")
        return redirect('restaurant:home')

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_profile':
            # Update profile information
            try:
                request.user.first_name = request.POST.get('first_name', '').strip()
                request.user.last_name = request.POST.get('last_name', '').strip()
                request.user.email = request.POST.get('email', '').strip()
                request.user.phone_number = request.POST.get('phone_number', '').strip()
                request.user.address = request.POST.get('address', '').strip()
                
                # Validation
                if not request.user.first_name or not request.user.last_name:
                    messages.error(request, "First name and last name are required.")
                    return redirect('admin_panel:profile')
                
                if request.user.email and User.objects.filter(email=request.user.email).exclude(id=request.user.id).exists():
                    messages.error(request, "Email already exists.")
                    return redirect('admin_panel:profile')
                
                request.user.save()
                messages.success(request, "Profile updated successfully.")
                
            except Exception as e:
                messages.error(request, f"Error updating profile: {str(e)}")
                
        elif action == 'change_password':
            # Change password
            try:
                current_password = request.POST.get('current_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')
                
                # Validation
                if not current_password or not new_password or not confirm_password:
                    messages.error(request, "All password fields are required.")
                    return redirect('admin_panel:profile')
                
                if not request.user.check_password(current_password):
                    messages.error(request, "Current password is incorrect.")
                    return redirect('admin_panel:profile')
                
                if new_password != confirm_password:
                    messages.error(request, "New passwords do not match.")
                    return redirect('admin_panel:profile')
                
                if len(new_password) < 8:
                    messages.error(request, "Password must be at least 8 characters long.")
                    return redirect('admin_panel:profile')
                
                request.user.set_password(new_password)
                request.user.save()
                messages.success(request, "Password changed successfully.")
                
            except Exception as e:
                messages.error(request, f"Error changing password: {str(e)}")
        return redirect('admin_panel:profile')
    
    context = {
        'user': request.user,
    }
    
    return render(request, 'admin_panel/profile.html', context)

@login_required
def manage_qr_code(request):
    """QR Code management for restaurant owners"""
    if not request.user.is_owner():
        messages.error(request, "Access denied. Owner privileges required.")
        return redirect('restaurant:home')
    
    # Ensure QR code exists
    if not request.user.restaurant_qr_code:
        request.user.generate_qr_code()
        request.user.save()
    
    # Generate the full QR URL using helper function
    qr_url = get_production_qr_url(request, request.user.restaurant_qr_code)
    
    context = {
        'user': request.user,
        'qr_code': request.user.restaurant_qr_code,
        'qr_url': qr_url,
        'restaurant_name': request.user.restaurant_name,
    }
    
    return render(request, 'admin_panel/manage_qr_code.html', context)

@login_required
@require_POST
def regenerate_qr_code(request):
    """Regenerate QR code for restaurant owner"""
    if not request.user.is_owner():
        messages.error(request, "Access denied. Owner privileges required.")
        return redirect('restaurant:home')
    
    # Generate new QR code
    import uuid
    request.user.restaurant_qr_code = f"REST-{uuid.uuid4().hex[:12].upper()}"
    request.user.save()
    
    messages.success(request, 'QR code has been regenerated successfully!')
    return redirect('admin_panel:manage_qr_code')

@login_required
def generate_qr_image(request):
    """Generate QR code image for restaurant owner"""
    if not request.user.is_owner():
        return HttpResponse("Access denied", status=403)
    
    # Ensure QR code exists
    if not request.user.restaurant_qr_code:
        request.user.generate_qr_code()
        request.user.save()
    
    # Generate the full QR URL using helper function
    qr_url = get_production_qr_url(request, request.user.restaurant_qr_code)
    
    # Create QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_url)
    qr.make(fit=True)

    # Create QR code image
    qr_img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to BytesIO
    img_io = io.BytesIO()
    qr_img.save(img_io, format='PNG')
    img_io.seek(0)
    
    # Return image response with NO CACHING to prevent stale QR codes
    response = HttpResponse(img_io.getvalue(), content_type='image/png')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
def debug_qr_code(request):
    """Debug endpoint to show QR code information"""
    if not request.user.is_owner():
        return HttpResponse("Access denied", status=403)
    
    from django.http import JsonResponse
    
    qr_url = get_production_qr_url(request, request.user.restaurant_qr_code)
    
    debug_info = {
        'username': request.user.username,
        'restaurant_name': request.user.restaurant_name,
        'qr_code_in_database': request.user.restaurant_qr_code,
        'full_qr_url': qr_url,
        'expected_access_url': f"/r/{request.user.restaurant_qr_code}/",
        'is_owner': request.user.is_owner(),
        'host': request.get_host(),
    }
    
    return JsonResponse(debug_info, json_dumps_params={'indent': 2})

@login_required
def import_products_csv(request):
    """Import products from CSV file"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    if request.method != 'POST':
        return redirect('admin_panel:manage_products')
    
    if 'csv_file' not in request.FILES:
        messages.error(request, 'Please select a CSV file to upload.')
        return redirect('admin_panel:manage_products')
    
    csv_file = request.FILES['csv_file']
    
    # Validate file type
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Please upload a valid CSV file.')
        return redirect('admin_panel:manage_products')
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        # Read CSV file
        decoded_file = csv_file.read().decode('utf-8-sig')  # utf-8-sig handles BOM
        csv_data = csv.DictReader(io.StringIO(decoded_file))
        
        imported_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_data, start=2):  # Start at 2 because row 1 is header
            try:
                # Validate required fields
                name = row.get('name', '').strip()
                price = row.get('price', '').strip()
                main_category_name = row.get('main_category', '').strip()
                
                if not name:
                    errors.append(f"Row {row_num}: Product name is required")
                    error_count += 1
                    continue
                
                if not price:
                    errors.append(f"Row {row_num}: Price is required")
                    error_count += 1
                    continue
                
                if not main_category_name:
                    errors.append(f"Row {row_num}: Main category is required")
                    error_count += 1
                    continue
                
                # Validate price
                try:
                    price_decimal = Decimal(str(price))
                    if price_decimal <= 0:
                        errors.append(f"Row {row_num}: Price must be greater than 0")
                        error_count += 1
                        continue
                except (InvalidOperation, ValueError):
                    errors.append(f"Row {row_num}: Invalid price format")
                    error_count += 1
                    continue
                
                # Find or create main category
                main_category_filter = {'name__iexact': main_category_name}
                if owner_filter:
                    main_category_filter['owner'] = owner_filter
                
                main_category = MainCategory.objects.filter(**main_category_filter).first()
                if not main_category:
                    main_category_data = {
                        'name': main_category_name,
                        'is_active': True,
                        'description': row.get('main_category_description', '').strip()
                    }
                    if owner_filter:
                        main_category_data['owner'] = owner_filter
                    main_category = MainCategory.objects.create(**main_category_data)
                    messages.info(request, f"Created main category '{main_category_name}' for this import.")
                
                # Handle subcategory
                sub_category = None
                sub_category_name = row.get('sub_category', '').strip()
                if sub_category_name:
                    sub_category = SubCategory.objects.filter(
                        name__iexact=sub_category_name,
                        main_category=main_category
                    ).first()
                    if not sub_category:
                        sub_category = SubCategory.objects.create(
                            main_category=main_category,
                            name=sub_category_name,
                            description=row.get('sub_category_description', '').strip(),
                            is_active=True
                        )
                        messages.info(request, f"Created sub category '{sub_category_name}' under '{main_category_name}'.")
                
                # Check if product already exists
                existing_product = Product.objects.filter(
                    name__iexact=name,
                    main_category=main_category
                ).first()
                
                if existing_product:
                    errors.append(f"Row {row_num}: Product '{name}' already exists in '{main_category_name}'")
                    error_count += 1
                    continue
                
                # Create product
                product_data = {
                    'name': name,
                    'description': row.get('description', '').strip(),
                    'main_category': main_category,
                    'sub_category': sub_category,
                    'price': price_decimal,
                    'available_in_stock': max(0, int(row.get('available_in_stock', 0) or 0)),
                    'is_available': str(row.get('is_available', 'true')).lower() in ['true', '1', 'yes', 'available'],
                    'preparation_time': max(1, int(row.get('preparation_time', 15) or 15)),
                }
                
                Product.objects.create(**product_data)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
                continue
        
        # Show results
        if imported_count > 0:
            messages.success(request, f'Successfully imported {imported_count} products!')
        
        if error_count > 0:
            error_message = f'{error_count} errors occurred during import:'
            if len(errors) <= 10:
                error_message += '\n' + '\n'.join(errors)
            else:
                error_message += '\n' + '\n'.join(errors[:10]) + f'\n... and {len(errors) - 10} more errors'
            messages.error(request, error_message)
        
        if imported_count == 0 and error_count == 0:
            messages.warning(request, 'No data found in the CSV file.')
            
    except Exception as e:
        messages.error(request, f'Error processing CSV file: {str(e)}')
    
    return redirect('admin_panel:manage_products')


@login_required
def import_products_excel(request):
    """Import products from Excel file"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    if request.method != 'POST':
        return redirect('admin_panel:manage_products')
    
    if 'excel_file' not in request.FILES:
        messages.error(request, 'Please select an Excel file to upload.')
        return redirect('admin_panel:manage_products')
    
    excel_file = request.FILES['excel_file']
    
    # Validate file type
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
        return redirect('admin_panel:manage_products')
    
    if not openpyxl:
        messages.error(request, 'Excel import is not available. Please contact administrator.')
        return redirect('admin_panel:manage_products')
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        # Save file temporarily
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        for chunk in excel_file.chunks():
            temp_file.write(chunk)
        temp_file.close()
        
        try:
            # Read Excel file
            workbook = openpyxl.load_workbook(temp_file.name, read_only=True, data_only=True)
            
            # Use first sheet or 'Products' sheet if exists
            sheet_name = 'Products' if 'Products' in workbook.sheetnames else workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
            
            # Get header row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).lower().strip())
                else:
                    headers.append('')
            
            # Map column indices
            col_mapping = {}
            for i, header in enumerate(headers):
                if 'name' in header and 'name' not in col_mapping:
                    col_mapping['name'] = i
                elif 'price' in header:
                    col_mapping['price'] = i
                elif 'category' in header and 'sub' not in header:
                    col_mapping['main_category'] = i
                elif 'sub' in header and 'category' in header:
                    col_mapping['sub_category'] = i
                elif 'description' in header:
                    col_mapping['description'] = i
                elif 'stock' in header or 'quantity' in header:
                    col_mapping['available_in_stock'] = i
                elif 'available' in header or 'status' in header:
                    col_mapping['is_available'] = i
                elif 'time' in header and 'prep' in header:
                    col_mapping['preparation_time'] = i
            
            if 'name' not in col_mapping or 'price' not in col_mapping or 'main_category' not in col_mapping:
                messages.error(request, 'Excel file must contain columns for Name, Price, and Main Category.')
                return redirect('admin_panel:manage_products')
            
            imported_count = 0
            error_count = 0
            errors = []
            
            # Process data rows
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # Skip empty rows
                        continue
                    
                    # Extract data
                    name = str(row[col_mapping['name']] or '').strip()
                    price = str(row[col_mapping['price']] or '').strip()
                    main_category_name = str(row[col_mapping['main_category']] or '').strip()
                    
                    if not name:
                        errors.append(f"Row {row_num}: Product name is required")
                        error_count += 1
                        continue
                    
                    if not price:
                        errors.append(f"Row {row_num}: Price is required")
                        error_count += 1
                        continue
                    
                    if not main_category_name:
                        errors.append(f"Row {row_num}: Main category is required")
                        error_count += 1
                        continue
                    
                    # Validate price
                    try:
                        price_decimal = Decimal(str(price))
                        if price_decimal <= 0:
                            errors.append(f"Row {row_num}: Price must be greater than 0")
                            error_count += 1
                            continue
                    except (InvalidOperation, ValueError):
                        errors.append(f"Row {row_num}: Invalid price format")
                        error_count += 1
                        continue
                    
                    # Find main category
                    main_category_filter = {'name__iexact': main_category_name}
                    if owner_filter:
                        main_category_filter['owner'] = owner_filter
                    
                    main_category = MainCategory.objects.filter(**main_category_filter).first()
                    if not main_category:
                        errors.append(f"Row {row_num}: Main category '{main_category_name}' not found")
                        error_count += 1
                        continue
                    
                    # Handle subcategory
                    sub_category = None
                    if 'sub_category' in col_mapping:
                        sub_category_name = str(row[col_mapping['sub_category']] or '').strip()
                        if sub_category_name:
                            sub_category = SubCategory.objects.filter(
                                name__iexact=sub_category_name,
                                main_category=main_category
                            ).first()
                            if not sub_category:
                                errors.append(f"Row {row_num}: Sub category '{sub_category_name}' not found")
                                error_count += 1
                                continue
                    
                    # Check if product exists
                    existing_product = Product.objects.filter(
                        name__iexact=name,
                        main_category=main_category
                    ).first()
                    
                    if existing_product:
                        errors.append(f"Row {row_num}: Product '{name}' already exists")
                        error_count += 1
                        continue
                    
                    # Create product
                    product_data = {
                        'name': name,
                        'description': str(row[col_mapping.get('description', 0)] or '').strip(),
                        'main_category': main_category,
                        'sub_category': sub_category,
                        'price': price_decimal,
                        'available_in_stock': max(0, int(row[col_mapping.get('available_in_stock', 0)] or 0)),
                        'preparation_time': max(1, int(row[col_mapping.get('preparation_time', 0)] or 15)),
                    }
                    
                    # Handle availability
                    if 'is_available' in col_mapping:
                        available_val = str(row[col_mapping['is_available']] or 'true').lower()
                        product_data['is_available'] = available_val in ['true', '1', 'yes', 'available']
                    else:
                        product_data['is_available'] = True
                    
                    Product.objects.create(**product_data)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
                    continue
            
            workbook.close()
            
            # Show results
            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} products from Excel!')
            
            if error_count > 0:
                error_message = f'{error_count} errors occurred during import:'
                if len(errors) <= 10:
                    error_message += '\n' + '\n'.join(errors)
                else:
                    error_message += '\n' + '\n'.join(errors[:10]) + f'\n... and {len(errors) - 10} more errors'
                messages.error(request, error_message)
            
            if imported_count == 0 and error_count == 0:
                messages.warning(request, 'No data found in the Excel file.')
                
        finally:
            # Clean up temp file
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
                
    except Exception as e:
        messages.error(request, f'Error processing Excel file: {str(e)}')
    
    return redirect('admin_panel:manage_products')


@login_required
def download_template_csv(request):
    """Download CSV template for product import"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="product_import_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'name',
        'description',
        'main_category',
        'sub_category',
        'price',
        'available_in_stock',
        'is_available',
        'preparation_time'
    ])
    
    # Add sample data
    writer.writerow([
        'Sample Pizza',
        'Delicious cheese pizza with fresh toppings',
        'Main Dishes',
        'Pizza',
        '12.99',
        '50',
        'true',
        '20'
    ])
    writer.writerow([
        'Sample Burger',
        'Juicy beef burger with lettuce and tomato',
        'Main Dishes',
        'Burgers',
        '8.99',
        '30',
        'true',
        '15'
    ])
    
    return response


@login_required
def download_template_excel(request):
    """Download Excel template for product import"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    if not openpyxl:
        messages.error(request, 'Excel export is not available. Please contact administrator.')
        return redirect('admin_panel:manage_products')
    
    # Create workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"
    
    # Headers
    headers = [
        'name',
        'description', 
        'main_category',
        'sub_category',
        'price',
        'available_in_stock',
        'is_available',
        'preparation_time'
    ]
    
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Sample data
    sample_data = [
        ['Sample Pizza', 'Delicious cheese pizza with fresh toppings', 'Main Dishes', 'Pizza', 12.99, 50, True, 20],
        ['Sample Burger', 'Juicy beef burger with lettuce and tomato', 'Main Dishes', 'Burgers', 8.99, 30, True, 15],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to BytesIO
    excel_io = io.BytesIO()
    workbook.save(excel_io)
    excel_io.seek(0)
    
    response = HttpResponse(
        excel_io.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="product_import_template.xlsx"'
    
    return response


@login_required
@require_POST
def bulk_delete_products(request):
    """Bulk delete multiple products"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'error': 'Access denied. Administrator or Owner privileges required.'})
    
    try:
        # Parse JSON data
        data = json.loads(request.body)
        product_ids = data.get('product_ids', [])
        
        if not product_ids:
            return JsonResponse({'success': False, 'error': 'No products selected for deletion.'})
        
        if len(product_ids) > 50:  # Limit bulk operations
            return JsonResponse({'success': False, 'error': 'Cannot delete more than 50 products at once.'})
        
        owner_filter = get_owner_filter(request.user)
        
        # Build query for products to delete
        if owner_filter:
            # Owner can only delete their own products
            products_to_delete = Product.objects.filter(
                id__in=product_ids,
                main_category__owner=owner_filter
            )
        else:
            # Administrator can delete all products
            products_to_delete = Product.objects.filter(id__in=product_ids)
        
        # Check if all requested products exist and are accessible
        found_count = products_to_delete.count()
        if found_count != len(product_ids):
            return JsonResponse({
                'success': False, 
                'error': f'Some products could not be found or you do not have permission to delete them. Found {found_count} out of {len(product_ids)} products.'
            })
        
        # Check for products that are in active orders (optional business logic)
        # You can uncomment this if you want to prevent deletion of products with active orders
        # from orders.models import OrderItem
        # active_order_products = OrderItem.objects.filter(
        #     product__in=products_to_delete,
        #     order__status__in=['pending', 'confirmed', 'preparing']
        # ).values_list('product_id', flat=True).distinct()
        # 
        # if active_order_products:
        #     return JsonResponse({
        #         'success': False,
        #         'error': f'Cannot delete products that are in active orders. {len(active_order_products)} products have active orders.'
        #     })
        
        # Get product names for logging
        product_names = list(products_to_delete.values_list('name', flat=True))
        
        # Perform bulk deletion
        deleted_count, deleted_details = products_to_delete.delete()
        
        # Log the deletion (optional)
        if hasattr(request.user, 'get_full_name'):
            user_name = request.user.get_full_name() or request.user.username
        else:
            user_name = request.user.username
            
        print(f"Bulk delete performed by {user_name}: {deleted_count} products deleted - {', '.join(product_names[:5])}")
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} products.'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data.'})
    except Exception as e:
        print(f"Error in bulk delete: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An error occurred while deleting products.'})
