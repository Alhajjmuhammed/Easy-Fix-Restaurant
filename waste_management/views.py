from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import models
from django.db.models import Q, Sum, Count, Avg
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth, TruncYear
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import datetime, timedelta, date
from decimal import Decimal
import json
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
from io import BytesIO

from accounts.models import get_owner_filter
from orders.models import Order, OrderItem
from restaurant.models import Product
from .models import FoodWasteLog, OrderCostBreakdown, WasteReportSummary, ProductCostSettings


def detect_automatic_waste():
    """Automatically detect potential waste scenarios"""
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    
    # Scenario 1: Orders that were never picked up (older than 2 hours)
    pickup_timeout = timezone.now() - timedelta(hours=2)
    abandoned_orders = Order.objects.filter(
        status='confirmed',
        created_at__lt=pickup_timeout,
    ).exclude(
        waste_logs__waste_reason='customer_left'  # Don't duplicate
    )
    
    auto_waste_count = 0
    
    for order in abandoned_orders:
        # Get a system user to record the waste
        system_user = User.objects.filter(role__name='administrator').first()
        if not system_user:
            system_user = User.objects.filter(role__name='owner').first()
        
        if not system_user:
            continue
        
        # Create waste logs for each item in the order
        for item in order.order_items.all():
            # Check if we already recorded waste for this item
            existing_waste = FoodWasteLog.objects.filter(
                order=order,
                product=item.product,
                waste_reason='customer_left'
            ).exists()
            
            if not existing_waste:
                waste_log = FoodWasteLog.objects.create(
                    order=order,
                    product=item.product,
                    quantity_wasted=item.quantity,
                    waste_reason='customer_left',
                    disposal_method='waste_bin',
                    notes=f'🤖 Auto-detected: Order #{order.id} abandoned after 2 hours',
                    recorded_by=system_user
                )
                
                # Calculate costs
                waste_log.calculate_costs()
                waste_log.save()
                
                auto_waste_count += 1
    
    return auto_waste_count


@login_required
def waste_dashboard(request):
    """Waste management dashboard with filtering and export functionality"""
    # Redirect cashiers to their specific waste recording page
    if request.user.is_cashier():
        return redirect('waste_management:cashier_record')
        
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied. Only owners and administrators can access waste management dashboard.")
        return redirect('restaurant:home')
    
    # Run automatic waste detection
    auto_detected = detect_automatic_waste()
    if auto_detected > 0:
        messages.info(request, f"🤖 Automatically detected {auto_detected} waste incidents from abandoned orders.")
    
    owner_filter = get_owner_filter(request.user)
    
    # Get filter parameters (similar to reports)
    period = request.GET.get('period', 'today')
    category_id = request.GET.get('category_id', 'all')
    subcategory_id = request.GET.get('subcategory_id', 'all')
    waste_reason = request.GET.get('waste_reason', 'all')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # Base queryset
    if request.user.is_administrator():
        waste_logs = FoodWasteLog.objects.all()
    else:
        waste_logs = FoodWasteLog.objects.filter(
            product__main_category__owner=owner_filter
        )
    
    # Apply period filter
    today = timezone.now().date()
    if period == 'today':
        waste_logs = waste_logs.filter(created_at__date=today)
    elif period == 'weekly':
        week_start = today - timedelta(days=today.weekday())
        waste_logs = waste_logs.filter(created_at__date__gte=week_start)
    elif period == 'monthly':
        month_start = today.replace(day=1)
        waste_logs = waste_logs.filter(created_at__date__gte=month_start)
    elif period == 'yearly':
        year_start = today.replace(month=1, day=1)
        waste_logs = waste_logs.filter(created_at__date__gte=year_start)
    
    # Apply category and subcategory filters
    if category_id != 'all':
        waste_logs = waste_logs.filter(product__main_category_id=category_id)
    
    if subcategory_id != 'all':
        waste_logs = waste_logs.filter(product__sub_category_id=subcategory_id)
    
    # Apply waste reason filter
    if waste_reason != 'all':
        waste_logs = waste_logs.filter(waste_reason=waste_reason)
    
    # Apply date filters (these override period if specified)
    if from_date:
        waste_logs = waste_logs.filter(created_at__date__gte=from_date)
    if to_date:
        waste_logs = waste_logs.filter(created_at__date__lte=to_date)
    
    # Calculate summary data
    total_items_wasted = waste_logs.aggregate(total=Sum('quantity_wasted'))['total'] or 0
    total_waste_cost = waste_logs.aggregate(total=Sum('total_cost'))['total'] or Decimal('0.00')
    avg_cost_per_item = (total_waste_cost / total_items_wasted) if total_items_wasted > 0 else 0
    
    # Waste by reason
    waste_by_reason = waste_logs.values('waste_reason').annotate(
        count=Count('id'),
        total_cost=Sum('total_cost')
    ).order_by('-total_cost')
    
    # Top wasted products
    top_wasted_products = waste_logs.values('product__name').annotate(
        total_quantity=Sum('quantity_wasted'),
        total_cost=Sum('total_cost')
    ).order_by('-total_cost')[:10]
    
    # Waste logs with pagination
    logs_list = waste_logs.select_related('product', 'order', 'recorded_by').order_by('-created_at')
    
    # Pagination
    page_number = request.GET.get('page', 1)
    paginator = Paginator(logs_list, 10)  # Same as reports (10 per page)
    page_obj = paginator.get_page(page_number)
    
    # Get categories and subcategories for filters
    from restaurant.models import MainCategory, SubCategory
    categories = MainCategory.objects.filter(owner=owner_filter)
    subcategories = SubCategory.objects.filter(main_category__owner=owner_filter)
    
    # Get products for the dropdown
    if request.user.is_administrator():
        products = Product.objects.filter(is_available=True).select_related('main_category', 'sub_category').order_by('name')
    else:
        products = Product.objects.filter(
            main_category__owner=owner_filter,
            is_available=True
        ).select_related('main_category', 'sub_category').order_by('name')
    
    # Filter subcategories by selected category if applicable
    if category_id != 'all':
        subcategories = subcategories.filter(main_category_id=category_id)
    
    # Waste reason choices
    waste_reason_choices = [
        ('overcooked', 'Overcooked'),
        ('expired', 'Expired'),
        ('customer_left', 'Customer Left'),
        ('preparation_error', 'Preparation Error'),
        ('contaminated', 'Contaminated'),
        ('spoiled', 'Spoiled'),
        ('other', 'Other'),
    ]
    
    context = {
        'total_items_wasted': total_items_wasted,
        'total_waste_cost': total_waste_cost,
        'avg_cost_per_item': avg_cost_per_item,
        'waste_by_reason': waste_by_reason,
        'top_wasted_products': top_wasted_products,
        'page_obj': page_obj,
        'waste_logs': page_obj,
        'categories': categories,
        'subcategories': subcategories,
        'products': products,  # Add products to context
        'waste_reason_choices': waste_reason_choices,
        'selected_category': category_id,
        'selected_subcategory': subcategory_id,
        'selected_period': period,
        'selected_waste_reason': waste_reason,
        'from_date': from_date,
        'to_date': to_date,
    }
    
    return render(request, 'waste_management/dashboard.html', context)


@login_required
def export_waste_csv(request):
    """Export waste data to CSV with same filtering as dashboard"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied.")
        return redirect('restaurant:home')
    
    # Get owner filter for multi-tenant support
    owner_filter = get_owner_filter(request.user)
    
    # Get same filters as dashboard
    period = request.GET.get('period', 'today')
    category_id = request.GET.get('category_id', 'all')
    subcategory_id = request.GET.get('subcategory_id', 'all')
    waste_reason = request.GET.get('waste_reason', 'all')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # Base queryset
    if request.user.is_administrator():
        waste_logs = FoodWasteLog.objects.all()
    else:
        waste_logs = FoodWasteLog.objects.filter(
            product__main_category__owner=owner_filter
        )
    
    # Apply same filters as dashboard
    today = timezone.now().date()
    if period == 'today':
        waste_logs = waste_logs.filter(created_at__date=today)
    elif period == 'weekly':
        week_start = today - timedelta(days=today.weekday())
        waste_logs = waste_logs.filter(created_at__date__gte=week_start)
    elif period == 'monthly':
        month_start = today.replace(day=1)
        waste_logs = waste_logs.filter(created_at__date__gte=month_start)
    elif period == 'yearly':
        year_start = today.replace(month=1, day=1)
        waste_logs = waste_logs.filter(created_at__date__gte=year_start)
    
    if category_id != 'all':
        waste_logs = waste_logs.filter(product__main_category_id=category_id)
    
    if subcategory_id != 'all':
        waste_logs = waste_logs.filter(product__sub_category_id=subcategory_id)
    
    if waste_reason != 'all':
        waste_logs = waste_logs.filter(waste_reason=waste_reason)
    
    if from_date:
        waste_logs = waste_logs.filter(created_at__date__gte=from_date)
    if to_date:
        waste_logs = waste_logs.filter(created_at__date__lte=to_date)
    
    # Calculate summary data
    total_items = waste_logs.aggregate(total=Sum('quantity_wasted'))['total'] or 0
    total_cost = waste_logs.aggregate(total=Sum('total_cost'))['total'] or Decimal('0.00')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="waste_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write summary
    writer.writerow(['Waste Management Report'])
    writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Period:', period.title()])
    writer.writerow(['Total Items Wasted:', total_items])
    writer.writerow(['Total Waste Cost:', f'${total_cost:.2f}'])
    writer.writerow([])  # Empty row
    
    # Write headers
    writer.writerow([
        'Date', 'Time', 'Product', 'Category', 'Sub Category', 
        'Quantity Wasted', 'Waste Reason', 'Total Cost', 
        'Ingredient Cost', 'Labor Cost', 'Overhead Cost',
        'Recorded By', 'Notes'
    ])
    
    # Write waste data
    for log in waste_logs.select_related('product', 'recorded_by').order_by('-created_at'):
        writer.writerow([
            log.created_at.strftime('%Y-%m-%d'),
            log.created_at.strftime('%H:%M:%S'),
            log.product.name,
            log.product.main_category.name,
            log.product.sub_category.name if log.product.sub_category else '',
            log.quantity_wasted,
            log.get_waste_reason_display(),
            f'${log.total_cost:.2f}',
            f'${log.ingredient_cost:.2f}',
            f'${log.labor_cost:.2f}',
            f'${log.overhead_cost:.2f}',
            log.recorded_by.get_full_name() if log.recorded_by else '',
            log.notes or ''
        ])
    
    return response


@login_required
def export_waste_pdf(request):
    """Export waste data to PDF with same filtering as dashboard"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied.")
        return redirect('restaurant:home')
    
    # Get same filtered data as CSV export (same filtering logic)
    owner_filter = get_owner_filter(request.user)
    period = request.GET.get('period', 'today')
    category_id = request.GET.get('category_id', 'all')
    subcategory_id = request.GET.get('subcategory_id', 'all')
    waste_reason = request.GET.get('waste_reason', 'all')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # Same filtering logic as CSV
    if request.user.is_administrator():
        waste_logs = FoodWasteLog.objects.all()
    else:
        waste_logs = FoodWasteLog.objects.filter(
            product__main_category__owner=owner_filter
        )
    
    # Apply filters (same as CSV)
    today = timezone.now().date()
    if period == 'today':
        waste_logs = waste_logs.filter(created_at__date=today)
    elif period == 'weekly':
        week_start = today - timedelta(days=today.weekday())
        waste_logs = waste_logs.filter(created_at__date__gte=week_start)
    elif period == 'monthly':
        month_start = today.replace(day=1)
        waste_logs = waste_logs.filter(created_at__date__gte=month_start)
    elif period == 'yearly':
        year_start = today.replace(month=1, day=1)
        waste_logs = waste_logs.filter(created_at__date__gte=year_start)
    
    if category_id != 'all':
        waste_logs = waste_logs.filter(product__main_category_id=category_id)
    if subcategory_id != 'all':
        waste_logs = waste_logs.filter(product__sub_category_id=subcategory_id)
    if waste_reason != 'all':
        waste_logs = waste_logs.filter(waste_reason=waste_reason)
    if from_date:
        waste_logs = waste_logs.filter(created_at__date__gte=from_date)
    if to_date:
        waste_logs = waste_logs.filter(created_at__date__lte=to_date)
    
    # Calculate summary
    total_items = waste_logs.aggregate(total=Sum('quantity_wasted'))['total'] or 0
    total_cost = waste_logs.aggregate(total=Sum('total_cost'))['total'] or Decimal('0.00')
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph("Waste Management Report", title_style))
    
    # Summary info
    summary_data = [
        ['Report Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Period:', period.title()],
        ['Total Items Wasted:', str(total_items)],
        ['Total Waste Cost:', f'${total_cost:.2f}'],
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Waste logs table
    if waste_logs.exists():
        # Table headers
        data = [['Date', 'Product', 'Category', 'Qty', 'Reason', 'Cost', 'Recorded By']]
        
        # Add waste data
        for log in waste_logs.select_related('product', 'recorded_by').order_by('-created_at'):
            data.append([
                log.created_at.strftime('%Y-%m-%d'),
                log.product.name[:20] + '...' if len(log.product.name) > 20 else log.product.name,
                log.product.main_category.name[:15] + '...' if len(log.product.main_category.name) > 15 else log.product.main_category.name,
                str(log.quantity_wasted),
                log.get_waste_reason_display()[:10] + '...' if len(log.get_waste_reason_display()) > 10 else log.get_waste_reason_display(),
                f'${log.total_cost:.2f}',
                (log.recorded_by.get_full_name()[:15] + '...' if len(log.recorded_by.get_full_name()) > 15 else log.recorded_by.get_full_name()) if log.recorded_by else ''
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    else:
        story.append(Paragraph("No waste records found for the selected criteria.", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="waste_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


@login_required
@require_http_methods(["POST"])
def auto_detect_waste(request):
    """API endpoint to manually trigger automatic waste detection"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        detected_count = detect_automatic_waste()
        return JsonResponse({
            'success': True,
            'detected_count': detected_count,
            'message': f'Auto-detection complete. Found {detected_count} waste incidents.'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def record_food_waste(request):
    """Record a food waste incident"""
    if not (request.user.is_administrator() or request.user.is_owner() or request.user.is_cashier()):
        return JsonResponse({'error': 'Access denied. Only owners, administrators, and cashiers can record waste.'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        # Get required data
        product_id = data.get('product_id')
        quantity_wasted = int(data.get('quantity_wasted', 0))
        waste_reason = data.get('waste_reason')
        disposal_method = data.get('disposal_method', 'waste_bin')
        notes = data.get('notes', '')
        
        if not product_id or quantity_wasted <= 0 or not waste_reason:
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        # Get product
        product = get_object_or_404(Product, id=product_id)
        
        # Create waste log
        waste_log = FoodWasteLog.objects.create(
            product=product,
            quantity_wasted=quantity_wasted,
            waste_reason=waste_reason,
            disposal_method=disposal_method,
            notes=notes,
            recorded_by=request.user
        )
        
        # Calculate costs
        waste_log.calculate_costs()
        waste_log.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Waste recorded successfully',
            'waste_log': {
                'id': waste_log.id,
                'product_name': waste_log.product.name,
                'quantity_wasted': waste_log.quantity_wasted,
                'total_cost': float(waste_log.total_cost),
                'waste_reason': waste_log.get_waste_reason_display(),
                'disposal_method': waste_log.get_disposal_method_display()
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
        
        if order_item_id:
            order_item = get_object_or_404(OrderItem, id=order_item_id)
            product = order_item.product
        elif product_id:
            product = get_object_or_404(Product, id=product_id)
        
        if not product:
            return JsonResponse({'error': 'Product is required'}, status=400)
        
        # Calculate costs
        cost_settings = ProductCostSettings.objects.filter(product=product).first()
        if cost_settings:
            ingredient_cost = cost_settings.ingredient_cost_per_unit * quantity_wasted
            labor_cost = cost_settings.labor_cost_per_unit * quantity_wasted
            overhead_cost = cost_settings.overhead_cost_per_unit * quantity_wasted
        else:
            # Default cost calculation (30% of menu price)
            base_cost = product.price * Decimal('0.30') * quantity_wasted
            ingredient_cost = base_cost * Decimal('0.60')
            labor_cost = base_cost * Decimal('0.25')
            overhead_cost = base_cost * Decimal('0.15')
        
        total_cost = ingredient_cost + labor_cost + overhead_cost
        
        # Create waste log
        waste_log = FoodWasteLog.objects.create(
            order=order,
            order_item=order_item,
            product=product,
            quantity_wasted=quantity_wasted,
            ingredient_cost=ingredient_cost,
            labor_cost=labor_cost,
            overhead_cost=overhead_cost,
            total_cost=total_cost,
            waste_reason=waste_reason,
            disposal_method=disposal_method,
            notes=notes,
            recorded_by=request.user
        )
        
        # Update order status if applicable
        if order and waste_reason in ['customer_refused', 'customer_left']:
            order.status = 'customer_refused'
            order.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Waste recorded: {quantity_wasted}x {product.name} - ${total_cost}',
            'waste_log_id': waste_log.id
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def waste_reports(request):
    """Waste reports with filters and export options"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied.")
        return redirect('restaurant:home')
    
    owner_filter = get_owner_filter(request.user)
    
    # Get filters
    period_type = request.GET.get('period', 'daily')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    waste_reason = request.GET.get('reason', '')
    product_filter = request.GET.get('product', '')
    
    # Set default dates based on period
    if not date_from or not date_to:
        today = timezone.now().date()
        if period_type == 'daily':
            date_from = today
            date_to = today
        elif period_type == 'weekly':
            date_from = today - timedelta(days=7)
            date_to = today
        elif period_type == 'monthly':
            date_from = today.replace(day=1)
            date_to = today
        elif period_type == 'yearly':
            date_from = today.replace(month=1, day=1)
            date_to = today
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Build queryset - simplified to avoid column ambiguity
    if request.user.is_administrator():
        waste_logs = FoodWasteLog.objects.all()
    else:
        waste_logs = FoodWasteLog.objects.filter(
            product__main_category__owner=owner_filter
        )
    
    # Apply filters
    waste_logs = waste_logs.filter(created_at__date__range=[date_from, date_to])
    
    if waste_reason:
        waste_logs = waste_logs.filter(waste_reason=waste_reason)
    
    if product_filter:
        waste_logs = waste_logs.filter(product__name__icontains=product_filter)
    
    # Generate report data
    report_data = generate_waste_report_data(waste_logs, period_type, date_from, date_to)
    
    # Get products for filter dropdown
    if request.user.is_administrator():
        products = Product.objects.all()
    else:
        products = Product.objects.filter(main_category__owner=owner_filter)
    
    # Get the actual waste records for the table
    waste_records = waste_logs.select_related('product', 'recorded_by').order_by('-created_at')
    
    context = {
        'report_data': report_data,
        'waste_records': waste_records,
        'period_type': period_type,
        'start_date': date_from.strftime('%Y-%m-%d') if date_from else '',
        'end_date': date_to.strftime('%Y-%m-%d') if date_to else '',
        'reason_filter': waste_reason,
        'product_filter': product_filter,
        'products': products,
        'waste_reason_choices': FoodWasteLog.WASTE_REASON_CHOICES,
    }
    
    return render(request, 'waste_management/reports.html', context)


def generate_waste_report_data(waste_logs, period_type, date_from, date_to):
    """Generate comprehensive report data"""
    
    # Summary statistics
    total_items = waste_logs.aggregate(total=Sum('quantity_wasted'))['total'] or 0
    total_cost = waste_logs.aggregate(total=Sum('total_cost'))['total'] or Decimal('0.00')
    total_ingredient_cost = waste_logs.aggregate(total=Sum('ingredient_cost'))['total'] or Decimal('0.00')
    total_labor_cost = waste_logs.aggregate(total=Sum('labor_cost'))['total'] or Decimal('0.00')
    total_overhead_cost = waste_logs.aggregate(total=Sum('overhead_cost'))['total'] or Decimal('0.00')
    
    # Breakdown by reason
    reason_breakdown = waste_logs.values('waste_reason').annotate(
        count=Count('id'),
        quantity=Sum('quantity_wasted'),
        cost=Sum('total_cost')
    ).order_by('-cost')
    
    # Breakdown by product
    product_breakdown = waste_logs.values('product__name', 'product__price').annotate(
        count=Count('id'),
        quantity=Sum('quantity_wasted'),
        cost=Sum('total_cost')
    ).order_by('-cost')
    
    # Breakdown by disposal method
    disposal_breakdown = waste_logs.values('disposal_method').annotate(
        count=Count('id'),
        quantity=Sum('quantity_wasted'),
        cost=Sum('total_cost')
    ).order_by('-cost')
    
    # Time series data
    if period_type == 'daily':
        time_series = waste_logs.annotate(
            period=TruncDate('created_at')
        ).values('period').annotate(
            items=Sum('quantity_wasted'),
            cost=Sum('total_cost')
        ).order_by('period')
    elif period_type == 'weekly':
        time_series = waste_logs.annotate(
            period=TruncWeek('created_at')
        ).values('period').annotate(
            items=Sum('quantity_wasted'),
            cost=Sum('total_cost')
        ).order_by('period')
    elif period_type == 'monthly':
        time_series = waste_logs.annotate(
            period=TruncMonth('created_at')
        ).values('period').annotate(
            items=Sum('quantity_wasted'),
            cost=Sum('total_cost')
        ).order_by('period')
    else:  # yearly
        time_series = waste_logs.annotate(
            period=TruncYear('created_at')
        ).values('period').annotate(
            items=Sum('quantity_wasted'),
            cost=Sum('total_cost')
        ).order_by('period')
    
    return {
        'summary': {
            'total_items': total_items,
            'total_cost': total_cost,
            'total_ingredient_cost': total_ingredient_cost,
            'total_labor_cost': total_labor_cost,
            'total_overhead_cost': total_overhead_cost,
            'average_cost_per_item': total_cost / total_items if total_items > 0 else Decimal('0.00'),
        },
        'reason_breakdown': list(reason_breakdown),
        'product_breakdown': list(product_breakdown),
        'disposal_breakdown': list(disposal_breakdown),
        'time_series': list(time_series),
        'waste_logs': waste_logs.select_related('product', 'order', 'recorded_by').order_by('-created_at')
    }


@login_required
def export_waste_report(request):
    """Export waste report in various formats"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    export_format = request.GET.get('format', 'csv')
    period_type = request.GET.get('period', 'daily')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    waste_reason = request.GET.get('reason', '')
    product_filter = request.GET.get('product', '')
    
    # Parse dates
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Get data
    owner_filter = get_owner_filter(request.user)
    
    if request.user.is_administrator():
        waste_logs = FoodWasteLog.objects.all()
    else:
        waste_logs = FoodWasteLog.objects.filter(
            product__main_category__owner=owner_filter
        )
    
    if date_from and date_to:
        waste_logs = waste_logs.filter(created_at__date__range=[date_from, date_to])
    
    if waste_reason:
        waste_logs = waste_logs.filter(waste_reason=waste_reason)
    
    if product_filter:
        waste_logs = waste_logs.filter(product__name__icontains=product_filter)
    
    # Generate report data
    report_data = generate_waste_report_data(waste_logs, period_type, date_from, date_to)
    
    # Export based on format
    if export_format == 'csv':
        return export_csv(report_data, date_from, date_to)
    elif export_format == 'excel':
        return export_excel(report_data, date_from, date_to)
    elif export_format == 'pdf':
        return export_pdf(report_data, date_from, date_to)
    else:
        return JsonResponse({'error': 'Invalid format'}, status=400)


def export_csv(report_data, date_from, date_to):
    """Export waste report as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="waste_report_{date_from}_to_{date_to}.csv"'
    
    writer = csv.writer(response)
    
    # Summary section
    writer.writerow(['WASTE REPORT SUMMARY'])
    writer.writerow(['Period', f'{date_from} to {date_to}'])
    writer.writerow(['Total Items Wasted', report_data['summary']['total_items']])
    writer.writerow(['Total Cost', f"${report_data['summary']['total_cost']:.2f}"])
    writer.writerow(['Average Cost per Item', f"${report_data['summary']['average_cost_per_item']:.2f}"])
    writer.writerow([])
    
    # Detailed logs
    writer.writerow(['DETAILED WASTE LOGS'])
    writer.writerow([
        'Date', 'Product', 'Quantity', 'Reason', 'Disposal Method',
        'Ingredient Cost', 'Labor Cost', 'Overhead Cost', 'Total Cost',
        'Order Number', 'Recorded By', 'Notes'
    ])
    
    for log in report_data['waste_logs']:
        writer.writerow([
            log.created_at.strftime('%Y-%m-%d %H:%M'),
            log.product.name,
            log.quantity_wasted,
            log.get_waste_reason_display(),
            log.get_disposal_method_display(),
            f"${log.ingredient_cost:.2f}",
            f"${log.labor_cost:.2f}",
            f"${log.overhead_cost:.2f}",
            f"${log.total_cost:.2f}",
            log.order.order_number if log.order else '',
            log.recorded_by.username,
            log.notes
        ])
    
    writer.writerow([])
    
    # Breakdown by reason
    writer.writerow(['WASTE BY REASON'])
    writer.writerow(['Reason', 'Count', 'Quantity', 'Total Cost'])
    for item in report_data['reason_breakdown']:
        writer.writerow([
            dict(FoodWasteLog.WASTE_REASON_CHOICES)[item['waste_reason']],
            item['count'],
            item['quantity'],
            f"${item['cost']:.2f}"
        ])
    
    writer.writerow([])
    
    # Breakdown by product
    writer.writerow(['WASTE BY PRODUCT'])
    writer.writerow(['Product', 'Count', 'Quantity', 'Total Cost'])
    for item in report_data['product_breakdown']:
        writer.writerow([
            item['product__name'],
            item['count'],
            item['quantity'],
            f"${item['cost']:.2f}"
        ])
    
    return response


def export_excel(report_data, date_from, date_to):
    """Export waste report as Excel"""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    
    # Summary sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    
    # Header styling
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Summary data
    summary_sheet['A1'] = 'WASTE REPORT SUMMARY'
    summary_sheet['A1'].font = header_font
    
    summary_sheet['A3'] = 'Period:'
    summary_sheet['B3'] = f'{date_from} to {date_to}'
    summary_sheet['A4'] = 'Total Items Wasted:'
    summary_sheet['B4'] = report_data['summary']['total_items']
    summary_sheet['A5'] = 'Total Cost:'
    summary_sheet['B5'] = float(report_data['summary']['total_cost'])
    summary_sheet['A6'] = 'Average Cost per Item:'
    summary_sheet['B6'] = float(report_data['summary']['average_cost_per_item'])
    
    # Format currency cells
    summary_sheet['B5'].number_format = '$#,##0.00'
    summary_sheet['B6'].number_format = '$#,##0.00'
    
    # Detailed logs sheet
    logs_sheet = workbook.create_sheet("Detailed Logs")
    headers = [
        'Date', 'Product', 'Quantity', 'Reason', 'Disposal Method',
        'Ingredient Cost', 'Labor Cost', 'Overhead Cost', 'Total Cost',
        'Order Number', 'Recorded By', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = logs_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row, log in enumerate(report_data['waste_logs'], 2):
        logs_sheet.cell(row=row, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M'))
        logs_sheet.cell(row=row, column=2, value=log.product.name)
        logs_sheet.cell(row=row, column=3, value=log.quantity_wasted)
        logs_sheet.cell(row=row, column=4, value=log.get_waste_reason_display())
        logs_sheet.cell(row=row, column=5, value=log.get_disposal_method_display())
        logs_sheet.cell(row=row, column=6, value=float(log.ingredient_cost))
        logs_sheet.cell(row=row, column=7, value=float(log.labor_cost))
        logs_sheet.cell(row=row, column=8, value=float(log.overhead_cost))
        logs_sheet.cell(row=row, column=9, value=float(log.total_cost))
        logs_sheet.cell(row=row, column=10, value=log.order.order_number if log.order else '')
        logs_sheet.cell(row=row, column=11, value=log.recorded_by.username)
        logs_sheet.cell(row=row, column=12, value=log.notes)
        
        # Format currency columns
        for col in [6, 7, 8, 9]:
            logs_sheet.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Auto-adjust column widths
    for sheet in [summary_sheet, logs_sheet]:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Reason breakdown sheet
    reason_sheet = workbook.create_sheet("Waste by Reason")
    reason_headers = ['Reason', 'Count', 'Quantity', 'Total Cost']
    
    for col, header in enumerate(reason_headers, 1):
        cell = reason_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row, item in enumerate(report_data['reason_breakdown'], 2):
        reason_sheet.cell(row=row, column=1, value=dict(FoodWasteLog.WASTE_REASON_CHOICES)[item['waste_reason']])
        reason_sheet.cell(row=row, column=2, value=item['count'])
        reason_sheet.cell(row=row, column=3, value=item['quantity'])
        reason_sheet.cell(row=row, column=4, value=float(item['cost']))
        reason_sheet.cell(row=row, column=4).number_format = '$#,##0.00'
    
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="waste_report_{date_from}_to_{date_to}.xlsx"'
    
    return response


def export_pdf(report_data, date_from, date_to):
    """Export waste report as PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.darkblue,
        alignment=1  # Center
    )
    story.append(Paragraph('FOOD WASTE REPORT', title_style))
    story.append(Spacer(1, 20))
    
    # Summary section
    summary_data = [
        ['Period:', f'{date_from} to {date_to}'],
        ['Total Items Wasted:', str(report_data['summary']['total_items'])],
        ['Total Cost:', f"${report_data['summary']['total_cost']:.2f}"],
        ['Average Cost per Item:', f"${report_data['summary']['average_cost_per_item']:.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Waste by reason
    story.append(Paragraph('Waste Breakdown by Reason', styles['Heading2']))
    reason_data = [['Reason', 'Count', 'Quantity', 'Total Cost']]
    
    for item in report_data['reason_breakdown'][:10]:  # Top 10
        reason_data.append([
            dict(FoodWasteLog.WASTE_REASON_CHOICES)[item['waste_reason']],
            str(item['count']),
            str(item['quantity']),
            f"${item['cost']:.2f}"
        ])
    
    reason_table = Table(reason_data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch])
    reason_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(reason_table)
    story.append(Spacer(1, 20))
    
    # Top wasted products
    story.append(Paragraph('Top Wasted Products', styles['Heading2']))
    product_data = [['Product', 'Count', 'Quantity', 'Total Cost']]
    
    for item in report_data['product_breakdown'][:10]:  # Top 10
        product_data.append([
            item['product__name'],
            str(item['count']),
            str(item['quantity']),
            f"${item['cost']:.2f}"
        ])
    
    product_table = Table(product_data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch])
    product_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(product_table)
    
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="waste_report_{date_from}_to_{date_to}.pdf"'
    
    return response


@login_required
def cost_settings(request):
    """Manage product cost settings"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        messages.error(request, "Access denied.")
        return redirect('restaurant:home')
    
    owner_filter = get_owner_filter(request.user)
    
    if request.user.is_administrator():
        products = Product.objects.all()
    else:
        products = Product.objects.filter(main_category__owner=owner_filter)
    
    # Get or create cost settings for each product
    cost_settings = []
    for product in products:
        setting, created = ProductCostSettings.objects.get_or_create(
            product=product,
            defaults={
                'ingredient_cost_per_unit': product.price * Decimal('0.18'),  # 18% of menu price
                'labor_cost_per_unit': product.price * Decimal('0.08'),       # 8% of menu price
                'overhead_cost_per_unit': product.price * Decimal('0.04'),    # 4% of menu price
            }
        )
        cost_settings.append(setting)
    
    context = {
        'cost_settings': cost_settings,
    }
    
    return render(request, 'waste_management/cost_settings.html', context)


@login_required
@require_http_methods(["POST"])
def update_cost_settings(request):
    """Update product cost settings"""
    if not (request.user.is_administrator() or request.user.is_owner()):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        setting_id = data.get('setting_id')
        
        setting = get_object_or_404(ProductCostSettings, id=setting_id)
        
        # Update costs
        setting.ingredient_cost_per_unit = Decimal(str(data.get('ingredient_cost', 0)))
        setting.labor_cost_per_unit = Decimal(str(data.get('labor_cost', 0)))
        setting.overhead_cost_per_unit = Decimal(str(data.get('overhead_cost', 0)))
        setting.preparation_time_minutes = int(data.get('prep_time', 0))
        setting.cooking_time_minutes = int(data.get('cook_time', 0))
        setting.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Cost settings updated for {setting.product.name}',
            'total_cost': float(setting.total_cost_per_unit)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def cashier_record_waste(request):
    """Simple waste recording page for cashiers - add only"""
    if not request.user.is_cashier():
        messages.error(request, "Access denied. This page is for cashiers only.")
        return redirect('restaurant:home')
    
    owner_filter = get_owner_filter(request.user)
    
    # Get products for the dropdown - filter by main_category__owner
    products = Product.objects.filter(main_category__owner=owner_filter)
    
    context = {
        'products': products,
    }
    
    return render(request, 'waste_management/cashier_record.html', context)


@login_required
def recent_waste_records(request):
    """API endpoint to get recent waste records for the current user"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=403)
    
    owner_filter = get_owner_filter(request.user)
    
    # Get recent records (last 10) for this user or restaurant
    waste_logs = FoodWasteLog.objects.filter(
        recorded_by=request.user,
        product__main_category__owner=owner_filter
    ).select_related('product').order_by('-created_at')[:10]
    
    # Format records for API response
    records = []
    for log in waste_logs:
        records.append({
            'id': log.id,
            'product_name': log.product.name,
            'quantity': log.quantity_wasted,
            'cost': float(log.total_cost),
            'reason': log.get_waste_reason_display(),
            'date': log.created_at.strftime('%b %d, %H:%M')
        })
    
    return JsonResponse({'records': records})


@login_required
def cashier_waste_form(request):
    """Enhanced cashier waste recording form with better UX"""
    if not request.user.is_cashier():
        messages.error(request, "Access denied. This page is for cashiers only.")
        return redirect('restaurant:home')
    
    owner_filter = get_owner_filter(request.user)
    
    # Get products for the dropdown - filter by main_category__owner
    products = Product.objects.filter(
        main_category__owner=owner_filter,
        is_available=True
    ).select_related('main_category').order_by('name')
    
    context = {
        'products': products,
    }
    
    return render(request, 'waste_management/cashier_waste_form.html', context)


@login_required
def cashier_waste_interface(request):
    """Completely separate cashier waste interface - NO ADMIN SIDEBAR"""
    if not request.user.is_cashier():
        messages.error(request, "Access denied. This page is for cashiers only.")
        return redirect('accounts:login')
    
    owner_filter = get_owner_filter(request.user)
    
    # Get products for the dropdown - filter by main_category__owner
    products = Product.objects.filter(
        main_category__owner=owner_filter,
        is_available=True
    ).select_related('main_category').order_by('name')
    
    context = {
        'products': products,
    }
    
    return render(request, 'waste_management/cashier_waste.html', context)